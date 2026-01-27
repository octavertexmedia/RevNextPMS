"""
Report Generation Functions
"""
import os
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Optional, Union
from django.conf import settings
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.models import Property, Inventory, RatePlan
from bookings.models import Reservation, Payment
from integrations.models import PropertyIntegration, SyncLog


def generate_revenue_report(tenant, start_date, end_date, format='pdf') -> Optional[BytesIO]:
    """Generate revenue report for a tenant"""
    payments = Payment.objects.filter(
        reservation__property__tenant=tenant,
        payment_status='COMPLETED',
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).select_related('reservation', 'reservation__property')
    
    # Calculate totals
    total_revenue = sum(p.amount.amount for p in payments)
    total_bookings = payments.count()
    
    # Group by property
    property_revenue = {}
    for payment in payments:
        prop_name = payment.reservation.property.name
        if prop_name not in property_revenue:
            property_revenue[prop_name] = {'revenue': Decimal('0'), 'bookings': 0}
        property_revenue[prop_name]['revenue'] += payment.amount.amount
        property_revenue[prop_name]['bookings'] += 1
    
    # Group by channel
    channel_revenue = {}
    for payment in payments:
        channel = payment.reservation.provider_name
        if channel not in channel_revenue:
            channel_revenue[channel] = {'revenue': Decimal('0'), 'bookings': 0}
        channel_revenue[channel]['revenue'] += payment.amount.amount
        channel_revenue[channel]['bookings'] += 1
    
    data = {
        'tenant': tenant.name,
        'period': f"{start_date} to {end_date}",
        'total_revenue': total_revenue,
        'total_bookings': total_bookings,
        'property_breakdown': property_revenue,
        'channel_breakdown': channel_revenue,
        'generated_at': timezone.now(),
    }
    
    if format == 'pdf':
        return generate_pdf_revenue_report(data)
    elif format == 'excel':
        return generate_excel_revenue_report(data)
    elif format == 'csv':
        return generate_csv_revenue_report(data)
    
    return None


def generate_occupancy_report(tenant, start_date, end_date, format='pdf') -> Optional[BytesIO]:
    """Generate occupancy report for a tenant"""
    inventories = Inventory.objects.filter(
        property__tenant=tenant,
        date__gte=start_date,
        date__lte=end_date
    ).select_related('property', 'room_type')
    
    # Calculate occupancy statistics
    total_room_nights = sum(inv.total_rooms for inv in inventories)
    occupied_room_nights = sum(
        inv.total_rooms - inv.available_rooms - inv.blocked_rooms 
        for inv in inventories
    )
    occupancy_rate = (occupied_room_nights / total_room_nights * 100) if total_room_nights > 0 else 0
    
    # Group by property
    property_occupancy = {}
    for inv in inventories:
        prop_name = inv.property.name
        if prop_name not in property_occupancy:
            property_occupancy[prop_name] = {
                'total_rooms': 0,
                'occupied_rooms': 0,
                'available_rooms': 0,
                'blocked_rooms': 0,
            }
        property_occupancy[prop_name]['total_rooms'] += inv.total_rooms
        property_occupancy[prop_name]['occupied_rooms'] += (inv.total_rooms - inv.available_rooms - inv.blocked_rooms)
        property_occupancy[prop_name]['available_rooms'] += inv.available_rooms
        property_occupancy[prop_name]['blocked_rooms'] += inv.blocked_rooms
    
    data = {
        'tenant': tenant.name,
        'period': f"{start_date} to {end_date}",
        'total_room_nights': total_room_nights,
        'occupied_room_nights': occupied_room_nights,
        'occupancy_rate': occupancy_rate,
        'property_breakdown': property_occupancy,
        'generated_at': timezone.now(),
    }
    
    if format == 'pdf':
        return generate_pdf_occupancy_report(data)
    elif format == 'excel':
        return generate_excel_occupancy_report(data)
    elif format == 'csv':
        return generate_csv_occupancy_report(data)
    
    return None


def generate_bookings_report(tenant, start_date, end_date, format='pdf') -> Optional[BytesIO]:
    """Generate bookings report for a tenant"""
    reservations = Reservation.objects.filter(
        property__tenant=tenant,
        check_in__gte=start_date,
        check_in__lte=end_date
    ).select_related('property', 'room_type')
    
    # Calculate statistics
    total_bookings = reservations.count()
    confirmed = reservations.filter(status='CONFIRMED').count()
    cancelled = reservations.filter(status='CANCELLED').count()
    pending = reservations.filter(status='PENDING').count()
    
    # Group by channel
    channel_bookings = {}
    for res in reservations:
        channel = res.provider_name
        if channel not in channel_bookings:
            channel_bookings[channel] = {
                'total': 0,
                'confirmed': 0,
                'cancelled': 0,
                'revenue': Decimal('0'),
            }
        channel_bookings[channel]['total'] += 1
        if res.status == 'CONFIRMED':
            channel_bookings[channel]['confirmed'] += 1
            channel_bookings[channel]['revenue'] += res.total_amount.amount
        elif res.status == 'CANCELLED':
            channel_bookings[channel]['cancelled'] += 1
    
    data = {
        'tenant': tenant.name,
        'period': f"{start_date} to {end_date}",
        'total_bookings': total_bookings,
        'confirmed': confirmed,
        'cancelled': cancelled,
        'pending': pending,
        'channel_breakdown': channel_bookings,
        'generated_at': timezone.now(),
    }
    
    if format == 'pdf':
        return generate_pdf_bookings_report(data)
    elif format == 'excel':
        return generate_excel_bookings_report(data)
    elif format == 'csv':
        return generate_csv_bookings_report(data)
    
    return None


def generate_channel_performance_report(tenant, start_date, end_date, format='pdf') -> Optional[BytesIO]:
    """Generate channel performance report"""
    integrations = PropertyIntegration.objects.filter(
        property__tenant=tenant,
        is_active=True
    ).select_related('platform', 'property')
    
    channel_stats = {}
    for integration in integrations:
        platform = integration.platform.name
        reservations = Reservation.objects.filter(
            property=integration.property,
            provider_name=platform,
            check_in__gte=start_date,
            check_in__lte=end_date
        )
        
        payments = Payment.objects.filter(
            reservation__in=reservations,
            payment_status='COMPLETED'
        )
        
        channel_stats[platform] = {
            'bookings': reservations.count(),
            'confirmed': reservations.filter(status='CONFIRMED').count(),
            'revenue': sum(p.amount.amount for p in payments),
            'avg_booking_value': sum(p.amount.amount for p in payments) / payments.count() if payments.count() > 0 else 0,
            'sync_success_rate': calculate_sync_success_rate(integration, start_date, end_date),
        }
    
    data = {
        'tenant': tenant.name,
        'period': f"{start_date} to {end_date}",
        'channels': channel_stats,
        'generated_at': timezone.now(),
    }
    
    if format == 'pdf':
        return generate_pdf_channel_performance_report(data)
    elif format == 'excel':
        return generate_excel_channel_performance_report(data)
    elif format == 'csv':
        return generate_csv_channel_performance_report(data)
    
    return None


def calculate_sync_success_rate(integration, start_date, end_date):
    """Calculate sync success rate for an integration"""
    syncs = SyncLog.objects.filter(
        property_integration=integration,
        created_at__gte=timezone.make_aware(datetime.combine(start_date, datetime.min.time())),
        created_at__lte=timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
    )
    
    if syncs.count() == 0:
        return 0
    
    successful = syncs.filter(status='SUCCESS').count()
    return (successful / syncs.count()) * 100


# PDF Generation Functions

def generate_pdf_revenue_report(data: Dict) -> BytesIO:
    """Generate PDF revenue report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER,
    )
    
    # Title
    story.append(Paragraph("Revenue Report", title_style))
    story.append(Spacer(1, 12))
    
    # Report Info
    info_data = [
        ['Tenant:', data['tenant']],
        ['Period:', data['period']],
        ['Generated At:', data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')],
    ]
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Summary
    summary_data = [
        ['Total Revenue', f"₹{data['total_revenue']:,.2f}"],
        ['Total Bookings', str(data['total_bookings'])],
    ]
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Property Breakdown
    story.append(Paragraph("Revenue by Property", styles['Heading2']))
    story.append(Spacer(1, 12))
    
    prop_data = [['Property', 'Revenue (₹)', 'Bookings']]
    for prop_name, stats in data['property_breakdown'].items():
        prop_data.append([prop_name, f"{stats['revenue']:,.2f}", str(stats['bookings'])])
    
    prop_table = Table(prop_data, colWidths=[3*inch, 2*inch, 1*inch])
    prop_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
    ]))
    story.append(prop_table)
    story.append(Spacer(1, 20))
    
    # Channel Breakdown
    story.append(Paragraph("Revenue by Channel", styles['Heading2']))
    story.append(Spacer(1, 12))
    
    channel_data = [['Channel', 'Revenue (₹)', 'Bookings']]
    for channel, stats in data['channel_breakdown'].items():
        channel_data.append([channel, f"{stats['revenue']:,.2f}", str(stats['bookings'])])
    
    channel_table = Table(channel_data, colWidths=[3*inch, 2*inch, 1*inch])
    channel_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
    ]))
    story.append(channel_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_pdf_occupancy_report(data: Dict) -> BytesIO:
    """Generate PDF occupancy report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER,
    )
    
    story.append(Paragraph("Occupancy Report", title_style))
    story.append(Spacer(1, 12))
    
    # Summary
    summary_data = [
        ['Total Room Nights', str(data['total_room_nights'])],
        ['Occupied Room Nights', str(data['occupied_room_nights'])],
        ['Occupancy Rate', f"{data['occupancy_rate']:.2f}%"],
    ]
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Property Breakdown
    prop_data = [['Property', 'Total Rooms', 'Occupied', 'Available', 'Blocked']]
    for prop_name, stats in data['property_breakdown'].items():
        prop_data.append([
            prop_name,
            str(stats['total_rooms']),
            str(stats['occupied_rooms']),
            str(stats['available_rooms']),
            str(stats['blocked_rooms']),
        ])
    
    prop_table = Table(prop_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
    prop_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(prop_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_pdf_bookings_report(data: Dict) -> BytesIO:
    """Generate PDF bookings report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER,
    )
    
    story.append(Paragraph("Bookings Report", title_style))
    story.append(Spacer(1, 12))
    
    # Summary
    summary_data = [
        ['Total Bookings', str(data['total_bookings'])],
        ['Confirmed', str(data['confirmed'])],
        ['Cancelled', str(data['cancelled'])],
        ['Pending', str(data['pending'])],
    ]
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Channel Breakdown
    channel_data = [['Channel', 'Total', 'Confirmed', 'Cancelled', 'Revenue (₹)']]
    for channel, stats in data['channel_breakdown'].items():
        channel_data.append([
            channel,
            str(stats['total']),
            str(stats['confirmed']),
            str(stats['cancelled']),
            f"{stats['revenue']:,.2f}",
        ])
    
    channel_table = Table(channel_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
    channel_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(channel_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_pdf_channel_performance_report(data: Dict) -> BytesIO:
    """Generate PDF channel performance report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER,
    )
    
    story.append(Paragraph("Channel Performance Report", title_style))
    story.append(Spacer(1, 12))
    
    # Channel Performance Table
    perf_data = [['Channel', 'Bookings', 'Confirmed', 'Revenue (₹)', 'Avg Value (₹)', 'Sync Success %']]
    for channel, stats in data['channels'].items():
        perf_data.append([
            channel,
            str(stats['bookings']),
            str(stats['confirmed']),
            f"{stats['revenue']:,.2f}",
            f"{stats['avg_booking_value']:,.2f}",
            f"{stats['sync_success_rate']:.2f}%",
        ])
    
    perf_table = Table(perf_data, colWidths=[2*inch, 1*inch, 1*inch, 1.5*inch, 1.5*inch, 1*inch])
    perf_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(perf_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


# Excel Generation Functions

def generate_excel_revenue_report(data: Dict) -> BytesIO:
    """Generate Excel revenue report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Report"
    
    # Header
    ws['A1'] = 'Revenue Report'
    ws['A1'].font = Font(size=16, bold=True, color='1e40af')
    ws.merge_cells('A1:D1')
    
    # Report Info
    ws['A3'] = 'Tenant:'
    ws['B3'] = data['tenant']
    ws['A4'] = 'Period:'
    ws['B4'] = data['period']
    ws['A5'] = 'Generated At:'
    ws['B5'] = data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')
    
    # Summary
    ws['A7'] = 'Total Revenue'
    ws['B7'] = f"₹{data['total_revenue']:,.2f}"
    ws['A8'] = 'Total Bookings'
    ws['B8'] = data['total_bookings']
    
    # Property Breakdown
    ws['A10'] = 'Revenue by Property'
    ws['A10'].font = Font(bold=True, size=12)
    ws['A11'] = 'Property'
    ws['B11'] = 'Revenue (₹)'
    ws['C11'] = 'Bookings'
    
    row = 12
    for prop_name, stats in data['property_breakdown'].items():
        ws[f'A{row}'] = prop_name
        ws[f'B{row}'] = float(stats['revenue'])
        ws[f'C{row}'] = stats['bookings']
        row += 1
    
    # Channel Breakdown
    ws[f'A{row+1}'] = 'Revenue by Channel'
    ws[f'A{row+1}'].font = Font(bold=True, size=12)
    ws[f'A{row+2}'] = 'Channel'
    ws[f'B{row+2}'] = 'Revenue (₹)'
    ws[f'C{row+2}'] = 'Bookings'
    
    row += 3
    for channel, stats in data['channel_breakdown'].items():
        ws[f'A{row}'] = channel
        ws[f'B{row}'] = float(stats['revenue'])
        ws[f'C{row}'] = stats['bookings']
        row += 1
    
    # Style headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ['A11', 'B11', 'C11', f'A{row-3}', f'B{row-3}', f'C{row-3}']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_excel_occupancy_report(data: Dict) -> BytesIO:
    """Generate Excel occupancy report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Occupancy Report"
    
    ws['A1'] = 'Occupancy Report'
    ws['A1'].font = Font(size=16, bold=True)
    
    ws['A3'] = 'Total Room Nights'
    ws['B3'] = data['total_room_nights']
    ws['A4'] = 'Occupied Room Nights'
    ws['B4'] = data['occupied_room_nights']
    ws['A5'] = 'Occupancy Rate'
    ws['B5'] = f"{data['occupancy_rate']:.2f}%"
    
    ws['A7'] = 'Property'
    ws['B7'] = 'Total Rooms'
    ws['C7'] = 'Occupied'
    ws['D7'] = 'Available'
    ws['E7'] = 'Blocked'
    
    row = 8
    for prop_name, stats in data['property_breakdown'].items():
        ws[f'A{row}'] = prop_name
        ws[f'B{row}'] = stats['total_rooms']
        ws[f'C{row}'] = stats['occupied_rooms']
        ws[f'D{row}'] = stats['available_rooms']
        ws[f'E{row}'] = stats['blocked_rooms']
        row += 1
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_excel_bookings_report(data: Dict) -> BytesIO:
    """Generate Excel bookings report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings Report"
    
    ws['A1'] = 'Bookings Report'
    ws['A1'].font = Font(size=16, bold=True)
    
    ws['A3'] = 'Total Bookings'
    ws['B3'] = data['total_bookings']
    ws['A4'] = 'Confirmed'
    ws['B4'] = data['confirmed']
    ws['A5'] = 'Cancelled'
    ws['B5'] = data['cancelled']
    ws['A6'] = 'Pending'
    ws['B6'] = data['pending']
    
    ws['A8'] = 'Channel'
    ws['B8'] = 'Total'
    ws['C8'] = 'Confirmed'
    ws['D8'] = 'Cancelled'
    ws['E8'] = 'Revenue (₹)'
    
    row = 9
    for channel, stats in data['channel_breakdown'].items():
        ws[f'A{row}'] = channel
        ws[f'B{row}'] = stats['total']
        ws[f'C{row}'] = stats['confirmed']
        ws[f'D{row}'] = stats['cancelled']
        ws[f'E{row}'] = float(stats['revenue'])
        row += 1
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_excel_channel_performance_report(data: Dict) -> BytesIO:
    """Generate Excel channel performance report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Channel Performance"
    
    ws['A1'] = 'Channel Performance Report'
    ws['A1'].font = Font(size=16, bold=True)
    
    ws['A3'] = 'Channel'
    ws['B3'] = 'Bookings'
    ws['C3'] = 'Confirmed'
    ws['D3'] = 'Revenue (₹)'
    ws['E3'] = 'Avg Value (₹)'
    ws['F3'] = 'Sync Success %'
    
    row = 4
    for channel, stats in data['channels'].items():
        ws[f'A{row}'] = channel
        ws[f'B{row}'] = stats['bookings']
        ws[f'C{row}'] = stats['confirmed']
        ws[f'D{row}'] = float(stats['revenue'])
        ws[f'E{row}'] = float(stats['avg_booking_value'])
        ws[f'F{row}'] = f"{stats['sync_success_rate']:.2f}%"
        row += 1
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# CSV Generation Functions

def generate_csv_revenue_report(data: Dict) -> BytesIO:
    """Generate CSV revenue report"""
    import csv
    buffer = BytesIO()
    writer = csv.writer(buffer)
    
    writer.writerow(['Revenue Report'])
    writer.writerow(['Tenant', data['tenant']])
    writer.writerow(['Period', data['period']])
    writer.writerow([])
    writer.writerow(['Total Revenue', f"₹{data['total_revenue']:,.2f}"])
    writer.writerow(['Total Bookings', data['total_bookings']])
    writer.writerow([])
    writer.writerow(['Property', 'Revenue (₹)', 'Bookings'])
    for prop_name, stats in data['property_breakdown'].items():
        writer.writerow([prop_name, f"{stats['revenue']:,.2f}", stats['bookings']])
    writer.writerow([])
    writer.writerow(['Channel', 'Revenue (₹)', 'Bookings'])
    for channel, stats in data['channel_breakdown'].items():
        writer.writerow([channel, f"{stats['revenue']:,.2f}", stats['bookings']])
    
    buffer.seek(0)
    return buffer


def generate_csv_occupancy_report(data: Dict) -> BytesIO:
    """Generate CSV occupancy report"""
    import csv
    buffer = BytesIO()
    writer = csv.writer(buffer)
    
    writer.writerow(['Occupancy Report'])
    writer.writerow(['Total Room Nights', data['total_room_nights']])
    writer.writerow(['Occupied Room Nights', data['occupied_room_nights']])
    writer.writerow(['Occupancy Rate', f"{data['occupancy_rate']:.2f}%"])
    writer.writerow([])
    writer.writerow(['Property', 'Total Rooms', 'Occupied', 'Available', 'Blocked'])
    for prop_name, stats in data['property_breakdown'].items():
        writer.writerow([
            prop_name,
            stats['total_rooms'],
            stats['occupied_rooms'],
            stats['available_rooms'],
            stats['blocked_rooms'],
        ])
    
    buffer.seek(0)
    return buffer


def generate_csv_bookings_report(data: Dict) -> BytesIO:
    """Generate CSV bookings report"""
    import csv
    buffer = BytesIO()
    writer = csv.writer(buffer)
    
    writer.writerow(['Bookings Report'])
    writer.writerow(['Total Bookings', data['total_bookings']])
    writer.writerow(['Confirmed', data['confirmed']])
    writer.writerow(['Cancelled', data['cancelled']])
    writer.writerow(['Pending', data['pending']])
    writer.writerow([])
    writer.writerow(['Channel', 'Total', 'Confirmed', 'Cancelled', 'Revenue (₹)'])
    for channel, stats in data['channel_breakdown'].items():
        writer.writerow([
            channel,
            stats['total'],
            stats['confirmed'],
            stats['cancelled'],
            f"{stats['revenue']:,.2f}",
        ])
    
    buffer.seek(0)
    return buffer


def generate_csv_channel_performance_report(data: Dict) -> BytesIO:
    """Generate CSV channel performance report"""
    import csv
    buffer = BytesIO()
    writer = csv.writer(buffer)
    
    writer.writerow(['Channel Performance Report'])
    writer.writerow(['Channel', 'Bookings', 'Confirmed', 'Revenue (₹)', 'Avg Value (₹)', 'Sync Success %'])
    for channel, stats in data['channels'].items():
        writer.writerow([
            channel,
            stats['bookings'],
            stats['confirmed'],
            f"{stats['revenue']:,.2f}",
            f"{stats['avg_booking_value']:,.2f}",
            f"{stats['sync_success_rate']:.2f}%",
        ])
    
    buffer.seek(0)
    return buffer
